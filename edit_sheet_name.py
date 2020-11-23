import os
from openpyxl import load_workbook
import time


def edit_sheet_name():
    pass
    return


if __name__ == '__main__':
    files = "E:\World of Warcraft\_classic_\WTF\Account\ZHAWLDX\SavedVariables\TradeSkillMaster.lua"
    id_name = "D:\MyBackup\Desktop\WOW怀旧服\商业\TSM\\nameB.txt"
    EXCEL_file_name = r"C:\Users\saiterlz.DESKTOP-ON65SQ2\PycharmProjects\untitled1\test.xlsx"
    wb = load_workbook(EXCEL_file_name)
    full_workbook = wb.get_sheet_names()  # 读取workbook所有的表名
    for i in full_workbook:
        print(i)
        table = wb.get_sheet_by_name(i)
        rows = table.max_row
        cols = table.max_column
        print("表sheet的列为:%s,行数为：%s" % (cols, rows))
        table.cell(1, 1).value = u"物品名称"
        table.column_dimensions["A"].width = 40
        table.cell(1, 2).value = u"最低价格"
        table.column_dimensions["B"].width = 10
        table.cell(1, 3).value = u"平均价格"
        table.column_dimensions["C"].width = 10
        table.cell(1, 4).value = u"拍卖数量"
        table.column_dimensions["D"].width = 10
        table.cell(1, 5).value = u"物品数量"
        table.column_dimensions["E"].width = 10
        table.cell(1, 6).value = u"TSM4最后更新数据时间"
        table.column_dimensions["F"].width = 25
        print("修改完成")
    wb.save(EXCEL_file_name)
