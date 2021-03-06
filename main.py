#!/usr/bin/python
# -*- coding: utf-8 -*-
# 本程序功能:读取由TSM插件命令/tsm scan扫描完后的AH所有端口信息,包含物品名称,最低价格,平均价格,当前拍卖量,扫描时间.等信息
# 通过本程序,生成一坐EXCEL表格来方便进行价格走势分析.
# from win32com.client import Dispatch
from win32com.client import Dispatch
import string
import json
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import numbers  # 数据格式
from openpyxl.styles import Alignment  # 对齐方式
from openpyxl.styles import Font  # 字体
from openpyxl.styles import PatternFill  # 导入填充模块
import time
import pymysql
import os
import configparser
from openpyxl.chart import (
    Series,
    LineChart,
    Reference,
)


def just_open(filename):
    abs_filename= os.path.abspath(filename)
    print(abs_filename)
    xlApp = Dispatch("Excel.Application")
    xlApp.Visible = False
    xlBook = xlApp.Workbooks.Open(abs_filename)
    xlBook.Save()
    xlBook.Close()


def id_to_name(filename):
    # id_name = os.path.abspath(filename)
    # print(id_name)
    ItemNames = {}
    with open(id_name,'r',encoding='utf8') as id_f:
        id_ret = id_f.readlines()
        # print(id_ret)
        for i in id_ret:
            arrStr = i.splitlines()
            # print(arrStr)
            if len(arrStr) > 0:
                for v in arrStr:
                    # print(v)
                    strI = v.split(":")
                    # print(type(strI))
                    arrI = strI
                    # print(arrI)
                    if len(arrI) == 2:
                        ItemNames[arrI[0]] = arrI[1]
            id_ret = id_f.readline()
    return ItemNames


def timestamp_datetime(value):
    if type(value) != int:
        value = int(value)
    format = '%Y-%m-%d %H:%M:%S'
    # value为传入的值为时间戳(整形)，如：1332888820
    value = time.localtime(value)
    ## 经过localtime转换后变成
    ## time.struct_time(tm_year=2012, tm_mon=3, tm_mday=28, tm_hour=6, tm_min=53, tm_sec=40, tm_wday=2, tm_yday=88, tm_isdst=0)
    # 最后再经过strftime函数转换为正常日期格式。
    dt = time.strftime(format, value)
    return dt


def date_style_transfomation(date, format_string1="%m-%d %H:%M:%S", format_string2="%m-%d %H-%M-%S"):
    time_array = time.strptime(date, format_string1)
    str_date = time.strftime(format_string2, time_array)
    return str_date


def to_db_value(file):  # 从程序 中拿 到数据
    sql_comm_list = []
    file = files
    with open(file, encoding='utf8') as f:
        ret = f.readline()
        while ret:
            ret = f.readline()
            if sprt_word in ret:
                idxName = ret.find("internalData@csvAuctionDBScan")
                # print('idxName=', idxName)
                subName = ret[5:idxName - 1]
                if subName:
                    if ret.find("lastScan"):
                        # "f@lliance - 比格沃斯@internalData@csvAuctionDBScan" 实例
                        # 格式化数据 ，例如：itemString,minBuyout,marketValue,numAuctions,quantity,lastScan\ni:14484,69000,69000,4,4,1605895492\n
                        idxStart = ret.find("lastScan")
                        subStr = ret[idxStart + 10:len(ret) - 3]
                        arrItems = subStr.split('\\n')
                        if arrItems != 0:
                            print('have data')  # 已找到需求的数据段
                            for tmp in arrItems:
                                # print('原始数据：',tmp)
                                sql_tmp = list(tmp.split(','))
                                ItemName = sql_tmp[0].split(":")
                                sql_tmp[0] = ItemName[1]
                                sql_tmp[5] = timestamp_datetime(sql_tmp[5])  # 处理时间
                                sql_tmp.append('0')
                                # print('sql数据：', sql_tmp)
                                # sql_comm = "insert into auction_history(item_id,min_price,ave_price,auction_num,quanlity,scan_time,is_del) values (%s,%s,%s,%s,%s,str_to_date(\'%s\','%%Y-%%m-%%d %%H:%%i:%%s'),%s);" % (sql_tmp[0], sql_tmp[1], sql_tmp[2], sql_tmp[3], sql_tmp[4], sql_tmp[5], sql_tmp[6])
                                # print('SQL语句',sql_comm)
                                sql_comm_list.append(sql_tmp)

    content = tuple(sql_comm_list)  # 批量写sql语句支持元组
    return content


def insert_to_db(file):  # 从程序 中拿 到数据
    conn = pymysql.connect("119.3.224.53", "root", "Test123abc", "wowclassic")
    cursor = conn.cursor()
    start = time.clock()
    sql_comm = "insert into auction_history(item_id,min_price,ave_price,auction_num,quanlity,scan_time,is_del) values (%s,%s,%s,%s,%s,%s,%s)"
    sql_comm_list = to_db_value(file)
    # print('insert_to_db',sql_comm_list)
    try:
        # 执行sql语句 executemany
        cursor.executemany(sql_comm, sql_comm_list)
        # 执行sql语句
        conn.commit()
    except pymysql.Error as e:
        # 发生错误时回滚
        print('执行sql出错，进行回滚', e)
        conn.rollback()
    conn.close()
    end = time.clock()
    print("executemany方法用时：", end - start, "秒")
    return print('处理写入到MYSQL')


# 给分析页添加新培加的sheet页的名字到A例第row+1行.
def add_sheet_name(workbook, dates):
    print(workbook, dates)
    ws = workbook.get_sheet_by_name("分析")  # 获取sheet分析这个对象
    print(ws.title)  # 验证是否正确访问这个sheet(分析）
    ws_rows_len = ws.max_row  # 行数
    ws_cols_len = ws.max_column  # 列数
    # print("读取本表的行数 %s 和列数 %s" % (ws_rows_len, ws_cols_len))
    ws.cell(row=ws_rows_len + 1, column=1).value = dates  # 将A列的日期写入到该单元格中，单元格中的内容 是用参数传递进来
    ws_rows_curent = ws_rows_len + 1  # 定位要写入的数据为当前得到的行数加1
    for i in range(2, ws_cols_len + 1):  # 开始 遍历写入单元格公式内容 ，遍历范围了列数加1，因为for循环的机制才加1。写入的数据是从第 2列开始
        this_col_name = ws.cell(row=1, column=i).value  # 验证当前表中第一行的字段值 是否存在
        if ws.cell(row=1, column=i).value != None:  # 通过ws.cell().value函数得到该 值 ，用来判断第 一行对应字段是否为None
            # 写入公式 =VLOOKUP(B$1,INDIRECT("'"&$A4&"'!A:H"),2,0)/10000
            #       "=VLOOKUP((B$1,INDIRECT("'" + dates + "'!A:H"),2,0)/10000 "
            col_letter_str = get_column_letter(i)  # 使用get_column_letter()函数得到列对应的字母，否则为数字，无法代入公式
            print("本列的物品为:%s 在 %s 列,从 %s 行,开始写入数据..." % (this_col_name, col_letter_str, ws_rows_curent))
            indirect_str = "A" + str(ws_rows_curent)  # 拼接excel 函数 INDIRECT()中表名的内容 前后要用&$表名&
            comm_strings = '=VLOOKUP(' + col_letter_str + '$1,INDIRECT("\'"&$' + indirect_str + '&"\'!A:H"),2,0)/10000'  # 将字符串拼接成为EXCEL公式，难度 ***** 五星
            # print(comm_strings)
            ws.cell(row=ws_rows_curent, column=i).value = comm_strings  # 将拼接好的公式 写入EXCEL表
            ws.cell(row=ws_rows_curent, column=i).number_format = '0.0000'  # 设置数据格式
            ws.cell(row=ws_rows_curent, column=i).alignment = Alignment(horizontal='right',
                                                                        vertical='center')  # 设置居中对齐
        else:
            break

# 开始按列找出最小值
def get_small_value_to_color(path_excel):
    wb = load_workbook(path_excel,data_only=True)
    ws = wb.get_sheet_by_name("分析")
    # 设置字体样式，设置字体为 微软雅黑，单下划线，颜色为蓝色,字体加粗
    yahei_font_u = Font(name=u'微软雅黑', underline='single', color='0000FF', bold=True)
    fille = PatternFill('solid', fgColor='c6efce')  # 设置填充颜色为 橙色
    def_fille = PatternFill('solid', fgColor='FFFFFF')  # 设置填充颜色为 白色
    print(ws.title)
    ws_rows_len = ws.max_row
    print('本 sheet 表一共有 %s 行(rows)' % ws_rows_len)
    ws_cols_len = ws.max_column
    print('本 sheet 表一共有 %s 列(columns)' % ws_cols_len)
    start_row = 4  # 定义起始行,EXCEL表中的数据列,从第4行开始
    for col in range(2, ws_cols_len + 1):  # 定位列
        temp_cell_value = float(10000000.0000)
        temp_cell_pos = []
        print('当前 是 第 %s 列.' % col)
        # col_str = get_column_letter(cols)
        # print(ws[col_str])
        for row in range(start_row, ws_rows_len + 1):  # 遍历方向是列,所以选择变更 值 为行的变化.进行循环
            # cells_value = ws.cell(row=rows, column=cols).value
            cells_value = ws.cell(row=row, column=col).value
            ws.cell(row, col).fill = def_fille  # 重置当前单元格的颜色,将以前着色的单元格恢复无底色
            ws.cell(row, col).number_format = '0.0000'  # 设置数据格式
            ws.cell(row, col).alignment = Alignment(horizontal='right')  # 设置居中对齐
            if cells_value == '#N/A' or cells_value == None:  # 判断单元格中的值 等于'#N/A ,无法使用,进行下一个循环

                print('当前 单元格的值 为:%s ,此值不可用! 当前单元格的坐标, 列为: %s -- 行为: %s' % (cells_value, col, row))
                continue
            elif cells_value == '#REF!' or cells_value == 0:
                print('当前 单元格的值 为:%s ,此值不可用! 当前单元格的坐标, 列为: %s -- 行为: %s' % (cells_value, col, row))
                continue
            else:
                cells_value = float(cells_value)
                print('当前 单元格的值 为:%s  , 当前单元格的坐标, 列为: %s -- 行为: %s' % (cells_value, col, row))
                if temp_cell_value > cells_value:
                    temp_cell_value = cells_value
                    temp_cell_pos = [row, col]
                    print('进行数据比较,结果是当前单元格的值 比较小.符合要求,数据为:%s ,数据的坐标为行%s ,列 %s ' % (
                        temp_cell_value, temp_cell_pos[0], temp_cell_pos[1]))
                    # ws.cell(temp_cell_pos[0], temp_cell_pos[1]).fill = fille
                    # ws.cell(row - 1, col).fill = def_fille
                elif temp_cell_value == cells_value:
                    temp_cell_pos = [row, col]
                    print('进行数据比较,结果是当前单元格的值 相等.例外,数据为:%s ,数据的坐标为行%s ,列 %s ' % (
                        temp_cell_value, temp_cell_pos[0], temp_cell_pos[1]))
                else:
                    print('进行数据比较,结果是当前 单元格的值 比较大.  不符合要求,数据为:', cells_value)
                    pass
        # ws.cell(temp_cell_pos[0],temp_cell_pos[1]).font=yahei_font_u
        ws.cell(temp_cell_pos[0], temp_cell_pos[1]).fill = fille
    print('比较大小着色完毕!进行保存')
    wb.save(path_excel)


def write_to_excel(files):
    file = files
    with open(file, encoding='utf8') as f:
        ret = f.readline()
        while ret:
            ret = f.readline()
            if sprt_word in ret:
                idxName = ret.find("internalData@csvAuctionDBScan")
                # print('idxName=', idxName)
                subName = ret[5:idxName - 1]
                if subName:
                    print('服务器文件名称为:', subName)
                    if ret.find("lastScan"):
                        # "f@lliance - 比格沃斯@internalData@csvAuctionDBScan" 实例
                        # 格式化数据 ，例如：itemString,minBuyout,marketValue,numAuctions,quantity,lastScan\ni:14484,69000,69000,4,4,1605895492\n
                        idxStart = ret.find("lastScan")
                        subStr = ret[idxStart + 10:len(ret) - 3]
                        arrItems = subStr.split('\\n')
                        if os.path.exists("%s.xlsx" % subName):
                            wb = load_workbook("%s.xlsx" % subName)
                        else:
                            wb = Workbook(data_only=True)
                        # AddSheet(fmt.Sprintf("%s", time.Now().Format("01-02 15-04-05"))
                        new_sheet_name = time.strftime("%m-%d %H-%M-%S", time.localtime())
                        ws = wb.create_sheet(new_sheet_name)
                        ws.cell(1, 1).value = u"物品名称"
                        ws.column_dimensions["B"].width = 20
                        # ws.row_dimensions[1].height = 40 #行高
                        ws.cell(1, 2).value = u"最低价格"
                        ws.column_dimensions["B"].width = 10
                        ws.cell(1, 3).value = u"平均价格"
                        ws.column_dimensions["C"].width = 10
                        ws.cell(1, 4).value = u"拍卖数量"
                        ws.column_dimensions["D"].width = 10
                        ws.cell(1, 5).value = u"物品数量"
                        ws.column_dimensions["E"].width = 10
                        ws.cell(1, 6).value = u"TSM4最后更新数据时间"
                        ws.column_dimensions["F"].width = 25
                        if arrItems != 0:
                            print('have data')  # 找到需求的数据段
                            for tmp in arrItems:
                                list_tmp = list(tmp.split(','))
                                ItemName = list_tmp[0].split(":")
                                list_tmp[0] = ItemNames[ItemName[1]]  # 处理名称
                                list_tmp[5] = timestamp_datetime(list_tmp[5])
                                ws.append(list_tmp)  # 写入数据到EXCEL
                        else:
                            print('no data ,error split data !')  # 没有找到需要数据段
                    else:
                        print('no data1')
                    add_sheet_name(wb, new_sheet_name)
                    wb.save("%s.xlsx" % subName)
    return print('处理写入到EXCEL')


if __name__ == "__main__":
    # 当前文件路径
    proDir = os.path.split(os.path.realpath(__file__))[0]
    # 在当前文件路径下查找.ini文件
    configPath = os.path.join(proDir, "config.ini")
    print(configPath)
    conf = configparser.ConfigParser()
    # 读取.ini文件
    conf.read(configPath, encoding="utf-8-sig")
    #控制单元
    open_write_to_excel_button = input('是否将-->拍卖行数据<--数据写入EXCEL,(1=开  0=关) :')
    compare_button = input('是否开启-->拍卖行最低价<--标注颜色,(1=开  0=关) :')
    open_to_sql_button = input('是否将-->拍卖行数据<--写入Mysql,(1=开  0=关) :')
    #原始设置部分
    # path_excel = "C:\\Users\sai\AppData\Local\Temp\TSM_Export_Excel.py\Alliance - 比格沃斯.xlsx"
    # Analysis_Sheet = "分析"
    # sprt_word = "csvAuctionDBScan"
    # files = "D:\World of Warcraft\_classic_\WTF\Account\ZHAWLDX\SavedVariables\TradeSkillMaster.lua"
    # id_name = "D:\\mystudy\\untitled1\\nameB.txt"
    #优化为读取INI文件，获取路径，路径是不能带有引号
    path_excel = conf.get('path', 'path_excel')
    files = conf.get('path', 'files')
    id_name = conf.get('path', 'id_name')

    Analysis_Sheet = conf.get('value', 'Analysis_Sheet')
    sprt_word = conf.get('value', 'sprt_word')
    ItemNames = id_to_name(id_name)
    # print(ItemNames)
    try:
        if open_to_sql_button != '0':
            insert_to_db(files)
        else:
            print('不用写入数据库')
    except Exception as err:
        print(err)

    try:
        if open_write_to_excel_button != "0":
            write_to_excel(files)
        else:
            print('不用写入EXCEL表')
    except Exception as err:
        print(err)
    try:
        if compare_button != "0":
            just_open(path_excel)
            get_small_value_to_color(path_excel)
        else:
            print('不用写入EXCEL表')
    except Exception as err:
        print(err)
    print(time.strftime('%Y-%m-%d %H:%M:%S',  time.localtime()))